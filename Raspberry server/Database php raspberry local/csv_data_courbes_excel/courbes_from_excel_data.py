'''Récupère les donnees de l'excel et trace des courbes avec les topics'''

import matplotlib.pyplot as plt
import openpyxl

PATH = '/home/thenes/database/SensorData.xlsx'

worksheet = openpyxl.load_workbook(PATH) #give the full path of the file here
sheet = worksheet.active


#listes vide
x = [] #temps
y1 = [] #capteur1
y2 = [] #capteur2
y3 = [] #capteur3

#ajoute les valeurs de chaque topic dans une liste
#on parcourt toutes les lignes de valeurs de la liste
for i in range(len(sheet['A'])-1):
    x.append(sheet.cell(row=i+2,column=7).value[11::])  #HEURE   -> date + heure : x.append(sheet.cell(row=i+2,column=7).value)
    y1.append(sheet.cell(row=i+2,column=4).value) #capteur1
    y2.append(sheet.cell(row=i+2,column=5).value) #capteur2
    y3.append(sheet.cell(row=i+2,column=6).value) #capteur3


plt.title("Capteurs en fonction du temps") 
plt.xlabel("temps") 
plt.ylabel("Capteurs")

plt.plot(x, y1, "r--", label="Capteur 1") # Récupérer nom dans l'excel : label=sheet['D1'].value)
plt.plot(x, y2, "b:o", label="Capteur 2")
plt.plot(x, y3, "g:o", label="Capteur 3") 
plt.legend()

plt.show()

