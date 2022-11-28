import paho.mqtt.client as mqtt
from openpyxl import Workbook
import time
from datetime import date, datetime


def on_connect(client, userdata, flags, rc):
    print("Connected with result code " + str(rc))

    # Subscribing in on_connect() means that if we lose the connection and
    # reconnect then subscriptions will be renewed.
    client.subscribe([("test", 2), ("test2", 1), ("topic3", 1)])





def on_message(client, userdata, message):
    print("Message received: " + message.topic + " : " + str(message.payload))

    workbook = Workbook()
    worksheet = workbook.active

    worksheet.title = 'Data IOT'

    topic_list = ["Date", "Heure", "Température", "Pollution"]
    topic=4

    # Noms des Topics
    for i in range (topic):
        worksheet.cell(row=1, column=i+1, value=topic_list[i])

    #enregistre les donnees des capteurs par ligne
    donnee_numero = 2 #commence à la deuxieme ligne pour les donnees
    nombre_donnees = 5

    while donnee_numero <= nombre_donnees + 2:
        today_date = date.today() #date aujourdhui

        now = datetime.now() 
        current_time = now.strftime("%H:%M:%S") #heure actuelle

        if message.topic == 'test':
            i = len(str(message.payload))
            temperature = str(message.payload)[2:i-1]



        #Date : colonne 1
        worksheet.cell(row=donnee_numero, column=1, value=today_date)

        #heure : colonne 2
        worksheet.cell(row=donnee_numero, column=2, value=current_time)

        #Topic 3 : colonne 
        worksheet.cell(row=donnee_numero, column=3, value=str(message.payload)) 

        #Topic 4 : colonne 4
        worksheet.cell(row=donnee_numero, column=4, value=25)   
        donnee_numero += 1

    #Enregistrement final
    workbook.save('/home/thenes/mqtt_connexion/tests/data.xlsx') #enregistre après changement







broker_address = "localhost"  # Broker address
port = 1883  # Broker port

#protocole TLS : securisé ----------- > https://www.frugalprototype.com/mqtt-tls/


# user = "yourUser"                    #Connection username
# password = "yourPassword"            #Connection password

client = mqtt.Client()  # create new instance
# client.username_pw_set(user, password=password)    #set username and password
client.on_connect = on_connect  # attach function to callback
client.on_message = on_message  # attach function to callback

client.connect(broker_address, port=port)  # connect to broker


client.loop_forever()