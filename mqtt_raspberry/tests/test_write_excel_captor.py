''' Creer fichier excel et écris sur l'excel : cree ligne de Topics et lignes avec les donnees des capteurs'''

from openpyxl import Workbook
from datetime import date, datetime
import time


workbook = Workbook()

worksheet = workbook.active


worksheet.title = 'Data IOT'

topic_list = ["Date", "Heure", "Température", "Pollution"]
topic=4

# Noms des Topics
for i in range (topic):
    worksheet.cell(row=1, column=i+1, value=topic_list[i])


#enregistre les donnees des capteurs par ligne

donnee_numero = 2 #commence à la deuxieme ligne pour les donnees
nombre_donnees = 5

while donnee_numero <= nombre_donnees + 2:
    today_date = date.today() #date aujourdhui

    now = datetime.now() 
    current_time = now.strftime("%H:%M:%S") #heure actuelle

    #Date : colonne 1
    worksheet.cell(row=donnee_numero, column=1, value=today_date)

    #heure : colonne 2
    worksheet.cell(row=donnee_numero, column=2, value=current_time)

    #Topic 3 : colonne 3
    worksheet.cell(row=donnee_numero, column=3, value=25)

    #Topic 4 : colonne 4
    worksheet.cell(row=donnee_numero, column=4, value=33.2)    

    donnee_numero += 1
    
    
    workbook.save('/home/thenes/mqtt_connexion/tests/data.xlsx') #enregistre après changement
    time.sleep(3)



#Enregistrement final
workbook.save('/home/thenes/mqtt_connexion/tests/data.xlsx') #enregistre après changement


