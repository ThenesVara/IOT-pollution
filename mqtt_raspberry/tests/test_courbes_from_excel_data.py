'''Récupère les donnees de l'excel et trace des courbes avec les topics'''

import numpy as np
import matplotlib.pyplot as plt
import openpyxl

worksheet = openpyxl.load_workbook('/home/thenes/mqtt_raspberry/tests/data.xlsx') #give the full path of the file here
sheet = worksheet.active

#nombre de lignes dans excel
#print(len(sheet['A']))

#nombre colonnes dans excel
#print(len(sheet['1']))

#listes vide
x = [] #temps
y1 = [] #capteur1
y2 = [] #capteur2

#ajoute les valeurs de chaque topic dans une liste
#on parcourt toutes les lignes de valeurs de la liste
for i in range(len(sheet['A'])-1):
    x.append(sheet.cell(row=i+2,column=2).value) 
    y1.append(sheet.cell(row=i+2,column=3).value) 
    y2.append(sheet.cell(row=i+2,column=4).value)


plt.xlabel(sheet['B1'].value) #temps
plt.ylabel("Capteur")

plt.plot(x, y1, "r--", label=sheet['C1'].value) #topic 3 : temperature
plt.plot(x, y2, "b:o", label=sheet['D1'].value) #topic 4 : pollution
plt.legend()

plt.show()

