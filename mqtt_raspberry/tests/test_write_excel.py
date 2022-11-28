'''Tuto ecrire sur un fichier libre office : https://www.youtube.com/watch?v=87aQ5DZlWlU'''
''' Creer un fichier excel et écris sur l'excel '''

from openpyxl import Workbook

workbook = Workbook()
workbook.save('/home/thenes/mqtt_connexion/tests/update.xlsx')

worksheet = workbook.active

print(worksheet)

worksheet.title = 'worksheet 1'
worksheet['A1'] = 0 #case en particulier vaut 0 : A colonne, 1 ligne
workbook.save('/home/thenes/mqtt_connexion/update.xlsx') #enregistre après changement


worksheet.cell(row=1, column=3, value=100)#logne 1, colonne 3, (C1) = 100
workbook.save('/home/thenes/mqtt_connexion/update.xlsx') #enregistre après changement



#creer autre worksheet
worksheet2 = workbook.create_sheet('worksheet2')
workbook.active = workbook['worksheet2'] #travaille sur l'autre worksheet
worksheet2['A1'] = 5

workbook.save('/home/thenes/mqtt_connexion/update.xlsx') #enregistre après changement

